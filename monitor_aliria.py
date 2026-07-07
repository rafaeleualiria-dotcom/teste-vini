# -*- coding: utf-8 -*-
"""
============================================================
 MONITOR ALIRIA — v3 (Vendas + Despesas + Liquidez)
============================================================
Monitora a planilha de controle e, a cada salvamento, envia
o dados.json atualizado para o repositório GitHub que
alimenta o dashboard (GitHub Pages).

O dados.json passa a conter 4 blocos:
  records        -> Vendas (26)                [mesmo formato de antes]
  despesas       -> Caixa Itau (25)            [lançamentos classificados]
  liquidezMensal -> Relatorio - 26 (Novo)      [Evolução do Ativo + Liquidez]
  liquidezDiaria -> Liquidez diária (Otimista) [projeção dia a dia]

Como usar:
  1. Ajuste as configurações abaixo (caminho da planilha e token).
  2. pip install openpyxl requests
  3. python monitor_aliria.py           -> fica monitorando
     python monitor_aliria.py --once    -> sincroniza uma vez e sai
============================================================
"""

import base64
import json
import os
import re
import sys
import time
import unicodedata
from datetime import datetime, date, timedelta

import requests
from openpyxl import load_workbook

# ============================================================
# CONFIGURAÇÕES
# ============================================================
XLSX_PATH = os.environ.get(
    "ALIRIA_XLSX",
    r"C:\Users\Vinicius\Documents\Aliria_Controle_Final_v1_-_2026_V4.xlsx",  # <- ajuste o caminho da SUA planilha
)
GH_OWNER = "SEU_USUARIO"      # <- seu usuário do GitHub
GH_REPO = "aliria-bi"         # <- nome do seu repositório
GH_FILE = "dados.json"
GH_BRANCH = "main"
# Token: por segurança, deixe no ambiente:  set GH_TOKEN=ghp_xxx  (Windows)
#                                           export GH_TOKEN=ghp_xxx (Mac/Linux)
GH_TOKEN = os.environ.get("GH_TOKEN", "")

POLL_SECONDS = 5  # intervalo de verificação de alteração do arquivo

# Abas
SHEET_VENDAS = "Vendas (26)"
SHEET_CAIXA = "Caixa Itau (25)"
SHEET_RELATORIO = "Relatorio - 26 (Novo)"
SHEET_LIQ_DIARIA = "Liquidez diária (Otimista)"

# Classificações do caixa que NÃO são despesa
# (compras de mercadoria, receitas e movimentações internas)
NAO_DESPESA = {
    "COMPRAS (VENDAS)",
    "RECEITA DE VENDAS",
    "RESGATE ITAU PRIVILEGE",
    "APLICACAO ITAU PRIVILEGE",
    "APP AUTOMATICA",
    "DEVOLUCAO VENDAS",
}

# ============================================================
# HELPERS
# ============================================================
EXCEL_EPOCH = date(1899, 12, 30)


def norm(s):
    """minúsculas, sem acento, espaços colapsados"""
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        if 20000 < v < 60000:  # serial Excel
            return EXCEL_EPOCH + timedelta(days=int(v))
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def to_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d,.\-]", "", str(v)).replace(".", "").replace(",", ".") \
        if ("," in str(v) and "." in str(v)) else re.sub(r"[^\d,.\-]", "", str(v)).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
            "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def mes_label(d):
    return f"{MESES_PT[d.month - 1]}/{d.year}" if d else ""


def ym(d):
    return f"{d.year}-{d.month:02d}" if d else ""


# ============================================================
# 1) VENDAS  — mesmo contrato do dados.json atual
# ============================================================
def ler_vendas(wb):
    ws = wb[SHEET_VENDAS]
    rows = list(ws.iter_rows(values_only=True))
    # localizar a linha de cabeçalho (a que contém "Fonte" e "DATA")
    hdr_idx = None
    for i, r in enumerate(rows[:20]):
        cells = [norm(c) for c in r]
        if "fonte" in cells and "data" in cells:
            hdr_idx = i
            break
    if hdr_idx is None:
        raise RuntimeError("Cabeçalho da aba Vendas não encontrado")

    recs = []
    for r in rows[hdr_idx + 1:]:
        if not r or r[0] in (None, ""):
            continue
        d = to_date(r[1])
        recs.append({
            "fonte": str(r[0]).strip(),
            "data": d.isoformat() if d else None,
            "mes": mes_label(d),
            "estado": (str(r[3]).strip() if r[3] else ""),
            "item": (str(r[4]).strip() if r[4] else ""),
            "laboratorio": (str(r[6]).strip() if r[6] else ""),
            "pagamento": (str(r[7]).strip() if r[7] is not None else ""),
            "precoCompra": to_num(r[8]),
            "totalCompra": to_num(r[9]),
            "icmsEnt": to_num(r[10]),
            "icmsSai": to_num(r[11]),
            "difal": to_num(r[12]),
            "irpj": to_num(r[13]),
            "cartao": to_num(r[14]),
            "cr": to_num(r[15]),
            "frete": to_num(r[16]),
            "precoVenda": to_num(r[17]),
            "qtd": to_num(r[18]),
            "valorTotal": to_num(r[19]),
            "margemVal": to_num(r[20]),
            "margemPct": to_num(r[21]),
            "difal2": to_num(r[22]) if len(r) > 22 else 0.0,
        })
    return recs


# ============================================================
# 2) DESPESAS — Caixa Itau (25)
# ============================================================
def ler_despesas(wb):
    ws = wb[SHEET_CAIXA]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [norm(c) for c in rows[0]]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    c_data = col("data")
    c_lanc = col("lancamento", "lançamento")
    c_val = col("valor (r$)", "valor")
    c_obs = col("obs")
    c_cls = col("classificacao", "classificação")

    despesas = []
    for r in rows[1:]:
        if not r or r[c_data] is None:
            continue
        cls = str(r[c_cls]).strip() if (c_cls is not None and r[c_cls]) else ""
        if not cls or cls.upper() in NAO_DESPESA:
            continue
        val = to_num(r[c_val])
        if val == 0:
            continue
        d = to_date(r[c_data])
        despesas.append({
            "data": d.isoformat() if d else None,
            "mes": ym(d),
            "mesLabel": mes_label(d),
            "categoria": cls,
            "descricao": (str(r[c_lanc]).strip() if c_lanc is not None and r[c_lanc] else ""),
            "obs": (str(r[c_obs]).strip() if c_obs is not None and r[c_obs] else ""),
            "valor": val,  # negativo = saída
        })
    return despesas


# ============================================================
# 3) LIQUIDEZ MENSAL — Relatorio - 26 (Novo)
#    Fórmula da própria planilha:
#    Liquidez = Caixa + Privilege DI + A Receber (todos)
#               - Depositado sem compra - A Pagar Fornecedores
#               - ICMS - IRPJ/CSLL - Difal - Doc a pagar
#    Liquidez com DIFAL = Liquidez + DIFAL Retido
# ============================================================
LIQ_LINHAS = {
    "caixa": "caixa",
    "privilege di": "privilegeDI",
    "a receber cr": "recCR",
    "a receber asaas": "recAsaas",
    "a receber licitacoes": "recLicitacoes",
    "a receber epharma": "recEpharma",
    "a receber unimed": "recUnimed",
    "a receber sami": "recSami",
    "a receber monte sinai": "recMonteSinai",
    "a receber alice": "recAlice",
    "depositado sem compra / n identificado": "depositadoSemCompra",
    "a pagar fornecedores": "pagarFornecedores",
    "icms": "icms",
    "irpj + csll": "irpjCsll",
    "difal a pagar": "difalPagar",
    "doc a pagar": "docPagar",
    "difal retido": "difalRetido",
    "estoque": "estoque",
    "loggo": "loggo",
    "liquidez": "liquidez",
    "liquidez com difal": "liquidezComDifal",
}


def ler_liquidez_mensal(wb):
    ws = wb[SHEET_RELATORIO]
    rows = list(ws.iter_rows(values_only=True))

    # linha de datas: a que contém "aliria" na coluna de rótulos (col idx 2)
    hdr_idx = None
    for i, r in enumerate(rows[:20]):
        if len(r) > 2 and norm(r[2]) == "aliria":
            hdr_idx = i
            break
    if hdr_idx is None:
        return None
    hdr = rows[hdr_idx]

    # mapear colunas -> mês (ignora colunas de total anual: 2023.0, 2024.0, "TOTAL")
    col_meses = []  # [(col_idx, 'YYYY-MM')]
    for c in range(3, len(hdr)):
        d = to_date(hdr[c])
        if d:
            col_meses.append((c, ym(d)))

    series = {}
    for r in rows:
        if len(r) > 2 and r[2] is not None:
            key = LIQ_LINHAS.get(norm(r[2]))
            # 'liquidez' pura só vale se a coluna de sinal (idx 1) for '='
            if key == "liquidez" and norm(r[1]) != "=":
                continue
            if key and key not in series:
                series[key] = [to_num(r[c]) if c < len(r) else 0.0 for c, _ in col_meses]
        # SG&A total mensal (linha "SG&A")
        if len(r) > 2 and norm(r[2]) == "sg&a" and "sga" not in series:
            series["sga"] = [to_num(r[c]) if c < len(r) else 0.0 for c, _ in col_meses]

    meses = [m for _, m in col_meses]

    # mantém só meses em que existe algum dado de liquidez
    liq = series.get("liquidez", [])
    ultimo = -1
    for i, v in enumerate(liq):
        # considera preenchido se algum componente relevante é != 0
        comp = abs(v) + abs(series.get("caixa", [0] * len(liq))[i]) + \
            abs(series.get("privilegeDI", [0] * len(liq))[i]) + \
            abs(series.get("pagarFornecedores", [0] * len(liq))[i])
        if comp > 0.005:
            ultimo = i
    if ultimo < 0:
        return None

    meses = meses[: ultimo + 1]
    series = {k: v[: ultimo + 1] for k, v in series.items()}
    return {"meses": meses, "series": series}


# ============================================================
# 4) LIQUIDEZ DIÁRIA — projeção da aba Liquidez diária (Otimista)
# ============================================================
def ler_liquidez_diaria(wb):
    ws = wb[SHEET_LIQ_DIARIA]
    rows = list(ws.iter_rows(values_only=True))

    fundos_saldo = 0.0
    hdr_idx = None
    for i, r in enumerate(rows[:10]):
        cells = [norm(c) for c in r]
        if "contas a pagar" in cells:
            hdr_idx = i
        for j, c in enumerate(cells):
            if c == "fundos + saldo" and i + 1 < len(rows):
                fundos_saldo = to_num(rows[i + 1][j])
    if hdr_idx is None:
        return None

    hdr = rows[hdr_idx]
    # colunas de recebíveis: da coluna 3 em diante, onde há rótulo
    rec_cols = [(c, str(hdr[c]).strip()) for c in range(3, len(hdr))
                if hdr[c] not in (None, "")]
    # remove duplicatas de rótulo mantendo a primeira ocorrência
    vistos, rec_cols_u = set(), []
    for c, nome in rec_cols:
        if nome not in vistos:
            vistos.add(nome)
            rec_cols_u.append((c, nome))

    dias = []
    for r in rows[hdr_idx + 1:]:
        d = to_date(r[0])
        if not d:
            continue
        receber = sum(to_num(r[c]) for c, _ in rec_cols_u if c < len(r))
        dias.append({
            "data": d.isoformat(),
            "pagar": to_num(r[1]),
            "receber": round(receber, 2),
            "saldo": to_num(r[2]),
        })
    if not dias:
        return None
    return {"fundosSaldo": fundos_saldo, "dias": dias}


# ============================================================
# GITHUB — envio do dados.json
# ============================================================
def push_github(payload):
    if not GH_TOKEN:
        raise RuntimeError("Defina o token: variável de ambiente GH_TOKEN")

    url = f"https://api.github.com/repos/{GH_OWNER}/{GH_REPO}/contents/{GH_FILE}"
    headers = {
        "Authorization": f"Bearer {GH_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }

    # sha atual (necessário para atualizar arquivo existente)
    sha = None
    r = requests.get(url, headers=headers, params={"ref": GH_BRANCH}, timeout=30)
    if r.status_code == 200:
        sha = r.json().get("sha")

    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    data = {
        "message": f"sync dados.json — {datetime.now():%d/%m/%Y %H:%M:%S}",
        "content": base64.b64encode(body.encode("utf-8")).decode("ascii"),
        "branch": GH_BRANCH,
    }
    if sha:
        data["sha"] = sha

    r = requests.put(url, headers=headers, json=data, timeout=60)
    r.raise_for_status()
    return len(body)


# ============================================================
# SINCRONIZAÇÃO
# ============================================================
def sincronizar():
    print(f"[{datetime.now():%H:%M:%S}] Lendo planilha…")
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    avisos = []

    records = ler_vendas(wb)
    print(f"  Vendas:   {len(records)} registros")

    try:
        despesas = ler_despesas(wb)
        print(f"  Despesas: {len(despesas)} lançamentos")
    except Exception as e:
        despesas = []
        avisos.append(f"Despesas: {e}")

    try:
        liq_mensal = ler_liquidez_mensal(wb)
        n = len(liq_mensal["meses"]) if liq_mensal else 0
        print(f"  Liquidez mensal: {n} meses")
    except Exception as e:
        liq_mensal = None
        avisos.append(f"Liquidez mensal: {e}")

    try:
        liq_diaria = ler_liquidez_diaria(wb)
        n = len(liq_diaria["dias"]) if liq_diaria else 0
        print(f"  Liquidez diária: {n} dias projetados")
    except Exception as e:
        liq_diaria = None
        avisos.append(f"Liquidez diária: {e}")

    wb.close()

    payload = {
        "savedAt": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z"),
        "version": 3,
        "records": records,
        "despesas": despesas,
        "liquidezMensal": liq_mensal,
        "liquidezDiaria": liq_diaria,
        "avisos": avisos,
    }

    tam = push_github(payload)
    print(f"[{datetime.now():%H:%M:%S}] ✅ Enviado ao GitHub ({tam/1024:.0f} KB)")


def main():
    if "--once" in sys.argv:
        sincronizar()
        return

    print("=" * 52)
    print(" MONITOR ALIRIA v3 — Vendas · Despesas · Liquidez")
    print("=" * 52)
    print(f" Planilha: {XLSX_PATH}")
    print(f" Destino:  github.com/{GH_OWNER}/{GH_REPO}/{GH_FILE}")
    print(" (Ctrl+C para encerrar)\n")

    last_mtime = 0
    while True:
        try:
            m = os.path.getmtime(XLSX_PATH)
            if m != last_mtime:
                if last_mtime:
                    time.sleep(2)  # aguarda o Excel terminar de gravar
                try:
                    sincronizar()
                    last_mtime = os.path.getmtime(XLSX_PATH)
                except PermissionError:
                    pass  # arquivo ainda aberto/gravando — tenta no próximo ciclo
                except Exception as e:
                    print(f"  ⚠️  Erro: {e}")
                    last_mtime = m
        except FileNotFoundError:
            print("  ⚠️  Planilha não encontrada — verifique o caminho.")
        time.sleep(POLL_SECONDS)


if __name__ == "__main__":
    main()
